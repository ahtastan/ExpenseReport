from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from sqlmodel import Session

from app.config import get_settings
from app.json_utils import decode_decimal
from app.models import ExpenseReport, ReportRun, StatementImport
from app.services import model_router
from app.services.receipt_annotations import ReceiptAnnotationLine, create_annotated_receipts_pdf
from app.services.report_validation import ReportValidation, validate_report_readiness
from app.services.review_sessions import confirmed_snapshot

_AMOUNT_QUANT = Decimal("0.0001")


DEFAULT_BUSINESS_REASONS = {
    date(2026, 3, 11): "Kartonsan Service Visit",
    date(2026, 3, 12): "Kartonsan Service Visit",
    date(2026, 3, 13): "Kartonsan Service Visit",
    date(2026, 4, 1): "Sanipak Visit",
}


REPORT_BUCKETS = [
    "Airfare/Bus/Ferry/Other",
    "Hotel/Lodging/Laundry",
    "Auto Rental",
    "Auto Gasoline",
    "Taxi/Parking/Tolls/Uber",
    "Other Travel Related",
    "Membership/Subscription Fees",
    "Customer Gifts",
    "Telephone/Internet",
    "Postage/Shipping",
    "Admin Supplies",
    "Lab Supplies",
    "Field Service Supplies",
    "Assets",
    "Other",
    "Meals/Snacks",
    "Breakfast",
    "Lunch",
    "Dinner",
    "Entertainment",
]


BUCKET_TOTAL_KEYS = {
    "airfare/bus/ferry/other": "airfare",
    "hotel/lodging/laundry": "hotel",
    "auto rental": "auto_rental",
    "auto gasoline": "gas",
    "taxi/parking/tolls/uber": "ground",
    "other travel related": "travel_other",
    "membership/subscription fees": "membership",
    "customer gifts": "gifts",
    "telephone/internet": "phone",
    "postage/shipping": "postage",
    "admin supplies": "admin",
    "lab supplies": "lab",
    "field service supplies": "field_service",
    "assets": "assets",
    "other": "other",
    "meals/snacks": "meal_m",
    "breakfast": "meal_b",
    "lunch": "meal_l",
    "dinner": "meal_d",
    "entertainment": "ent",
}


MEAL_DETAIL_CODES = {
    "meal_m": "M",
    "meal_b": "B",
    "meal_l": "L",
    "meal_d": "D",
    "ent": "E",
}


def _bucket_key(bucket: str) -> str:
    return " ".join(bucket.lower().replace("(", "").replace(")", "").split())


AIRFARE_BUCKET = "Airfare/Bus/Ferry/Other"

# Row layout for the AIR TRAVEL RECONCILIATION section per sheet.
# Week 1A: title=44, headers=45-46, data=47-49.
# Week 2A: title=45, headers=46-47, data=48-50.
AIR_TRAVEL_ROWS_BY_SHEET = {
    "Week 1A": [47, 48, 49],
    "Week 2A": [48, 49, 50],
}


def is_real_flight_line(line: "ReportLine") -> bool:
    """Discriminator: should this line populate the AIR TRAVEL RECONCILIATION block?

    The "Airfare/Bus/Ferry/Other" bucket name conflates real flights (which
    need ticket reconciliation — airline, RT/oneway, ticket cost, prior
    value) with ground transit (bus, ferry — daily totals only). The
    template's reconciliation block is flight-specific; bus/ferry rows
    landing here render as half-empty entries with no airline / RT-oneway /
    ticket-cost columns and confuse EDT auditors.

    Use operator-confirmed flight metadata as the marker: a line counts as
    a real flight only when bucket matches Airfare AND (airline name is
    non-empty OR an explicit total_tkt_cost is set). Buses/ferries share
    the bucket name but lack both signals; they land in row 7 daily totals
    via _allocate's day["airfare"] path and stay out of the reconciliation
    block.
    """
    if _bucket_key(line.report_bucket) != _bucket_key(AIRFARE_BUCKET):
        return False
    if (line.air_travel_airline or "").strip():
        return True
    if line.air_travel_total_tkt_cost is not None:
        return True
    return False


@dataclass(frozen=True)
class MealDetailLine:
    tx_date: date
    code: str
    place: str
    location: str
    participants: str
    reason: str
    amount: Decimal
    eg: bool
    mr: bool


@dataclass(frozen=True)
class ReportLine:
    transaction_id: int
    review_row_id: int | None
    receipt_id: int | None
    receipt_path: str | None
    receipt_file_name: str
    transaction_date: date
    supplier: str
    amount: Decimal
    currency: str
    business_or_personal: str
    report_bucket: str
    business_reason: str
    attendees: str
    air_travel_date: date | None = None
    air_travel_from: str | None = None
    air_travel_to: str | None = None
    air_travel_airline: str | None = None
    air_travel_rt_or_oneway: str | None = None
    air_travel_return_date: date | None = None
    air_travel_paid_by: str | None = None
    air_travel_total_tkt_cost: Decimal | None = None
    air_travel_prior_tkt_value: Decimal | None = None
    air_travel_comments: str | None = None
    meal_place: str | None = None
    meal_location: str | None = None
    meal_eg: bool = False
    meal_mr: bool = False


def _parse_optional_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def _parse_optional_decimal(value: object) -> Decimal | None:
    """Tolerantly decode an amount field from a confirmed_json blob.

    Accepts new string-shaped values (Decimal-as-string per M1 Day 2.5) and
    legacy float/int values written before the migration. Always quantizes
    to the 4-dp money grid so downstream arithmetic is on a common scale.
    """
    if value in (None, ""):
        return None
    try:
        decoded = decode_decimal(value)
    except (TypeError, InvalidOperation, ValueError):
        return None
    if decoded is None:
        return None
    return decoded.quantize(_AMOUNT_QUANT)


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "eg", "mr"}


def _confirmed_lines(
    session: Session, *, expense_report_id: int
) -> list[ReportLine]:
    _review, snapshot = confirmed_snapshot(session, expense_report_id=expense_report_id)
    lines: list[ReportLine] = []
    for row in snapshot:
        tx_date_raw = row.get("transaction_date")
        amount = row.get("amount")
        if not tx_date_raw or amount is None:
            continue
        tx_date = date.fromisoformat(str(tx_date_raw))
        lines.append(
            ReportLine(
                transaction_id=int(row["transaction_id"]),
                review_row_id=int(row["review_row_id"]) if row.get("review_row_id") is not None else None,
                receipt_id=int(row["receipt_id"]) if row.get("receipt_id") is not None else None,
                receipt_path=row.get("receipt_path"),
                receipt_file_name=row.get("receipt_file_name") or f"receipt_{row.get('receipt_id')}",
                transaction_date=tx_date,
                supplier=row.get("supplier") or "",
                amount=_parse_optional_decimal(amount) or Decimal("0"),
                currency=row.get("currency") or "TRY",
                business_or_personal=row.get("business_or_personal") or "",
                report_bucket=row.get("report_bucket") or "",
                business_reason=row.get("business_reason") or "",
                attendees=row.get("attendees") or "",
                air_travel_date=_parse_optional_date(row.get("air_travel_date")) or tx_date,
                air_travel_from=row.get("air_travel_from") or None,
                air_travel_to=row.get("air_travel_to") or None,
                air_travel_airline=row.get("air_travel_airline") or None,
                air_travel_rt_or_oneway=row.get("air_travel_rt_or_oneway") or None,
                air_travel_return_date=_parse_optional_date(row.get("air_travel_return_date")),
                air_travel_paid_by=row.get("air_travel_paid_by") or None,
                air_travel_total_tkt_cost=_parse_optional_decimal(row.get("air_travel_total_tkt_cost")),
                air_travel_prior_tkt_value=_parse_optional_decimal(row.get("air_travel_prior_tkt_value")),
                air_travel_comments=row.get("air_travel_comments") or None,
                meal_place=row.get("meal_place") or None,
                meal_location=row.get("meal_location") or None,
                meal_eg=_parse_bool(row.get("meal_eg")),
                meal_mr=_parse_bool(row.get("meal_mr")),
            )
        )
    return sorted(lines, key=lambda line: (line.transaction_date, line.review_row_id or 0))


def _allocate(line: ReportLine, day_totals: dict[date, dict[str, list[Decimal]]], detail_lines: dict[date, list[MealDetailLine]]) -> None:
    bp = line.business_or_personal.lower()
    day = day_totals[line.transaction_date]
    if bp != "business":
        day["other"].append(line.amount)
        return

    bucket_key = _bucket_key(line.report_bucket)
    total_key = BUCKET_TOTAL_KEYS.get(bucket_key, "other")
    total_amount = line.amount
    if bucket_key == _bucket_key(AIRFARE_BUCKET) and line.air_travel_total_tkt_cost is not None:
        total_amount = line.air_travel_total_tkt_cost
    day[total_key].append(total_amount)
    if total_key in MEAL_DETAIL_CODES:
        detail_lines[line.transaction_date].append(
            MealDetailLine(
                tx_date=line.transaction_date,
                code=MEAL_DETAIL_CODES[total_key],
                place=line.meal_place or line.supplier,
                location=line.meal_location or "",
                participants=line.attendees,
                reason=line.business_reason,
                amount=total_amount,
                eg=line.meal_eg,
                mr=line.meal_mr,
            )
        )


def group_meal_details_for_irs(
    details: list["MealDetailLine"],
) -> list[tuple["MealDetailLine", list[Decimal] | None]]:
    """Group same-supplier-same-code meal details for the Page 1B IRS section.

    Returns a list of (primary, sum_components) tuples in the input's
    first-occurrence order:

    - When a (code, normalized_supplier) appears exactly once → (detail, None).
      Caller writes the row normally; the amount cell inherits the template's
      pre-existing IF formula that pulls from the Week 1A daily total.

    - When the same (code, normalized_supplier) appears 2+ times → returns
      (first_detail, [d1.amount, d2.amount, …]). Caller writes the row using
      the first detail's metadata (place / location / participants / reason —
      identical across the duplicates by definition; this is the
      "split-bill on one dinner across two card transactions" pattern) and
      writes a SUM formula directly to the amount cell so the auditor can
      see =A+B and decode the component receipts.

    Different suppliers under the same code on the same day are kept as
    separate group entries — caller still has to allocate them to rows
    and currently drops 2nd/3rd different-supplier collisions per the
    template's 5-rows-per-day constraint. The Bug 2 fix targets the
    same-supplier collapse case explicitly; the rare different-supplier
    same-code collision is out of scope here.
    """
    groups: dict[tuple[str, str], list[MealDetailLine]] = {}
    for detail in details:
        supplier_norm = (detail.place or "").strip().lower()
        groups.setdefault((detail.code, supplier_norm), []).append(detail)
    result: list[tuple[MealDetailLine, list[Decimal] | None]] = []
    for group in groups.values():
        primary = group[0]
        if len(group) >= 2:
            result.append((primary, [d.amount for d in group]))
        else:
            result.append((primary, None))
    return result


def _resolve_period_ending(
    statement_date: date | None,
    lines: list[ReportLine],
) -> date | None:
    """Pick the period-ending date for the report header.

    Prefers the BMO statement_date when known (Diners Club statement-driven
    reports always have one). Falls back to the latest transaction_date in
    confirmed_lines when the statement_date is unset (manual entry path,
    or a statement uploaded before the statement-date column existed).
    Returns None only when neither is available, in which case callers
    leave the template's existing period-ending formula in place.
    """
    if statement_date is not None:
        return statement_date
    line_dates = [line.transaction_date for line in lines if line.transaction_date]
    return max(line_dates) if line_dates else None


def _apply_period_ending(
    wb,
    period_ending: date | None,
    *,
    has_week2_data: bool,
) -> None:
    """Override the template's period-ending chain with a concrete date.

    Template behavior we're replacing: ``Week 1A!M3 = ='Week 2A'!K5`` and
    ``Week 2A!M3 = =K5`` chain through a 14-column date projection that
    drifts past the actual data. For our 5-date November dataset the
    chain ends up at 2025-10-29 even though the latest transaction is
    2025-10-20 and the BMO statement closes on 2025-11-10.

    Override behavior:
      - Always overwrite ``Week 1A!M3`` with the resolved date. The B-side
        sheets (Week 1B, Week 2B) read M3 via IF wrappers; they pick up
        the overwrite for free.
      - When Week 2 has no data: also clear Week 2A's row-5 date formulas
        (E5–K5) and overwrite Week 2A!M3 with the same resolved date so
        the auditor doesn't see a stale projected period on a blank
        Week 2A.
      - When Week 2 HAS data: leave Week 2A's row-5 formulas alone (they
        derive valid Week 2 dates) and write the resolved date to Week 2A
        M3 too so its self-reference (=K5) is overridden.
    """
    if period_ending is None:
        return

    period_dt = datetime(period_ending.year, period_ending.month, period_ending.day)
    wb["Week 1A"]["M3"] = period_dt
    wb["Week 2A"]["M3"] = period_dt

    if not has_week2_data:
        # Clear Week 2A row-5 date formulas so the empty Week 2A doesn't
        # display projected dates 2025-10-23..29 next to a date label.
        # The cells stay empty; Excel renders blanks not zeros.
        for col in ("E", "F", "G", "H", "I", "J", "K"):
            wb["Week 2A"][f"{col}5"] = None


def _fill_workbook(
    template_path: Path,
    output_path: Path,
    employee_name: str,
    title: str,
    lines: list[ReportLine],
    *,
    period_ending: date | None = None,
) -> None:
    wb = load_workbook(template_path)
    ws1a, ws1b, ws2a, ws2b = wb["Week 1A"], wb["Week 1B"], wb["Week 2A"], wb["Week 2B"]
    ws1a["B3"] = employee_name
    ws1a["G3"] = title

    day_totals: dict[date, dict[str, list[Decimal]]] = defaultdict(lambda: defaultdict(list))
    detail_lines: dict[date, list[MealDetailLine]] = defaultdict(list)
    for line in lines:
        _allocate(line, day_totals, detail_lines)

    dates = sorted({line.transaction_date for line in lines})
    first7, next7 = dates[:7], dates[7:14]
    cols = ["E", "F", "G", "H", "I", "J", "K"]

    def fill_a(ws, page_dates: list[date]) -> None:
        def total_value(amounts: list[Decimal] | None) -> Decimal | str | None:
            if not amounts:
                return None
            if len(amounts) == 1:
                return amounts[0]
            parts = [f"{amount:.2f}".rstrip("0").rstrip(".") for amount in amounts]
            return "=" + "+".join(parts)

        for col, tx_date in zip(cols, page_dates):
            vals = day_totals[tx_date]
            ws[f"{col}5"] = datetime(tx_date.year, tx_date.month, tx_date.day)
            ws[f"{col}6"] = DEFAULT_BUSINESS_REASONS.get(tx_date, "")
            ws[f"{col}7"] = total_value(vals.get("airfare"))
            ws[f"{col}8"] = total_value(vals.get("hotel"))
            ws[f"{col}9"] = total_value(vals.get("auto_rental"))
            ws[f"{col}10"] = total_value(vals.get("gas"))
            ws[f"{col}11"] = total_value(vals.get("ground"))
            ws[f"{col}14"] = total_value(vals.get("travel_other"))
            ws[f"{col}18"] = total_value(vals.get("membership"))
            ws[f"{col}19"] = total_value(vals.get("gifts"))
            ws[f"{col}20"] = total_value(vals.get("phone"))
            ws[f"{col}21"] = total_value(vals.get("postage"))
            ws[f"{col}22"] = total_value(vals.get("admin"))
            ws[f"{col}23"] = total_value(vals.get("lab"))
            ws[f"{col}24"] = total_value(vals.get("field_service"))
            ws[f"{col}25"] = total_value(vals.get("assets"))
            ws[f"{col}26"] = total_value(vals.get("other"))
            ws[f"{col}29"] = total_value(vals.get("meal_m"))
            ws[f"{col}30"] = total_value(vals.get("meal_b"))
            ws[f"{col}31"] = total_value(vals.get("meal_l"))
            ws[f"{col}32"] = total_value(vals.get("meal_d"))
            ws[f"{col}35"] = total_value(vals.get("ent"))

    def fill_air_travel(ws, sheet_name: str, page_lines: list[ReportLine]) -> None:
        def travel_date_value(ln: ReportLine) -> datetime | str | None:
            travel_date = ln.air_travel_date or ln.transaction_date
            if travel_date is None:
                return None
            if (ln.air_travel_rt_or_oneway or "").strip().upper() == "RT" and ln.air_travel_return_date:
                return f"{travel_date:%d.%m.%Y} - {ln.air_travel_return_date:%d.%m.%Y}"
            return datetime(travel_date.year, travel_date.month, travel_date.day)

        rows = AIR_TRAVEL_ROWS_BY_SHEET.get(sheet_name, [])
        if not rows:
            return
        air_lines = [ln for ln in page_lines if is_real_flight_line(ln)]
        for row_num, ln in zip(rows, air_lines):
            value = travel_date_value(ln)
            if value is not None:
                ws[f"B{row_num}"] = value
            if ln.air_travel_from:
                ws[f"C{row_num}"] = ln.air_travel_from
            if ln.air_travel_to:
                ws[f"D{row_num}"] = ln.air_travel_to
            if ln.air_travel_airline:
                ws[f"E{row_num}"] = ln.air_travel_airline
            if ln.air_travel_rt_or_oneway:
                ws[f"F{row_num}"] = ln.air_travel_rt_or_oneway
            ws[f"G{row_num}"] = ln.air_travel_paid_by or "DC Card"
            if ln.air_travel_total_tkt_cost is not None:
                ws[f"H{row_num}"] = ln.air_travel_total_tkt_cost
            else:
                # Fall back to the line amount so the formula in column J has a value.
                ws[f"H{row_num}"] = ln.amount
            ws[f"I{row_num}"] = ln.air_travel_prior_tkt_value if ln.air_travel_prior_tkt_value is not None else 0
            # Column J holds the `=H-I` formula in the template — do not overwrite.
            if ln.air_travel_comments:
                ws[f"K{row_num}"] = ln.air_travel_comments

    def fill_b(ws, page_dates: list[date]) -> None:
        code_rows = {"M": 0, "B": 1, "L": 2, "D": 3, "E": 4}
        for day_index, tx_date in enumerate(page_dates):
            # Bug 2: collapse same-supplier-same-code duplicates into one IRS
            # row. The amount cell becomes a SUM formula so the auditor sees
            # each receipt's component (e.g. =92.72+16.65 for a dinner billed
            # across two card transactions). Without this, the second receipt
            # was silently dropped because it collided on (day, code).
            grouped = group_meal_details_for_irs(detail_lines.get(tx_date, []))
            used_codes: set[str] = set()
            for primary, sum_components in grouped:
                if primary.code not in code_rows or primary.code in used_codes:
                    continue
                used_codes.add(primary.code)
                rownum = 8 + day_index * 5 + code_rows[primary.code]
                ws[f"C{rownum}"] = primary.place[:40]
                ws[f"D{rownum}"] = primary.location[:28]
                ws[f"E{rownum}"] = primary.participants[:40]
                ws[f"F{rownum}"] = primary.reason[:50]
                ws[f"H{rownum}"] = "x" if primary.eg else None
                ws[f"I{rownum}"] = "x" if primary.mr else None
                # When 2+ receipts collapsed: write SUM formula directly to
                # the amount column (J) so the auditor sees =A+B and can
                # decode the component receipts. Without this, J{rownum}
                # inherits the template's pre-existing IF formula that
                # opaquely pulls the Week 1A daily total — which already
                # includes the sum, but doesn't reveal the components on
                # Page 1B itself.
                if sum_components is not None:
                    parts = [f"{amt:.2f}".rstrip("0").rstrip(".") for amt in sum_components]
                    ws[f"J{rownum}"] = "=" + "+".join(parts)

    first7_set = set(first7)
    next7_set = set(next7)
    first7_lines = [ln for ln in lines if ln.transaction_date in first7_set]
    next7_lines = [ln for ln in lines if ln.transaction_date in next7_set]

    fill_a(ws1a, first7)
    fill_a(ws2a, next7)
    fill_b(ws1b, first7)
    fill_b(ws2b, next7)
    fill_air_travel(ws1a, "Week 1A", first7_lines)
    fill_air_travel(ws2a, "Week 2A", next7_lines)
    _apply_period_ending(wb, period_ending, has_week2_data=bool(next7))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_summary(path: Path, validation: ReportValidation, lines: list[ReportLine], workbook_paths: list[Path]) -> None:
    summary_lines = [
        f"statement_import_id={validation.statement_import_id}",
        f"ready={validation.ready}",
        f"errors={validation.issue_count}",
        f"warnings={validation.warning_count}",
        f"approved_matches={validation.approved_matches}",
        f"report_lines={len(lines)}",
        f"business_receipts={validation.business_receipts}",
        f"personal_receipts={validation.personal_receipts}",
        "workbooks=" + ", ".join(p.name for p in workbook_paths),
        "",
        "issues:",
    ]
    summary_lines.extend(f"{issue.severity}|{issue.code}|{issue.message}" for issue in validation.issues)
    path.write_text("\n".join(summary_lines), encoding="utf-8")


def _bucket_totals(lines: list[ReportLine]) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for line in lines:
        amount = line.air_travel_total_tkt_cost if _bucket_key(line.report_bucket) == _bucket_key(AIRFARE_BUCKET) and line.air_travel_total_tkt_cost is not None else line.amount
        totals[line.report_bucket or "Other"] += amount
    return dict(sorted(totals.items()))


def _summary_payload(
    validation: ReportValidation,
    lines: list[ReportLine],
    workbook_paths: list[Path],
    employee_name: str,
    title_prefix: str,
) -> dict:
    business_reasons = sorted({line.business_reason for line in lines if line.business_reason})
    anomalies = [
        {
            "severity": issue.severity,
            "code": issue.code,
            "message": issue.message,
            "review_row_id": issue.review_row_id,
            "supplier": issue.supplier,
            "transaction_date": issue.transaction_date,
        }
        for issue in validation.issues
    ]
    anomalies.extend(
        {
            "severity": "warning",
            "code": "missing_receipt",
            "message": "Report line has no attached receipt file.",
            "transaction_id": line.transaction_id,
            "supplier": line.supplier,
            "transaction_date": line.transaction_date.isoformat(),
        }
        for line in lines
        if line.receipt_id is None or not line.receipt_path
    )
    return {
        "statement_import_id": validation.statement_import_id,
        "employee_name": employee_name,
        "title_prefix": title_prefix,
        "date_range": {
            "start": min((line.transaction_date for line in lines), default=None),
            "end": max((line.transaction_date for line in lines), default=None),
        },
        "trip_purpose_candidates": business_reasons,
        "totals_by_bucket": _bucket_totals(lines),
        "currency": sorted({line.currency for line in lines if line.currency}),
        "line_count": len(lines),
        "workbooks": [path.name for path in workbook_paths],
        "anomalies": anomalies,
    }


def _fallback_summary_markdown(payload: dict) -> str:
    purposes = payload.get("trip_purpose_candidates") or []
    totals = payload.get("totals_by_bucket") or {}
    anomalies = payload.get("anomalies") or []
    currencies = ", ".join(payload.get("currency") or [])
    lines = [
        "# Expense Report Summary",
        "",
        f"Trip purpose: {', '.join(purposes) if purposes else 'Not specified'}.",
        "",
        "Totals by bucket:",
    ]
    if totals:
        lines.extend(f"- {bucket}: {amount:.2f}{(' ' + currencies) if currencies else ''}" for bucket, amount in totals.items())
    else:
        lines.append("- No report lines.")
    lines.extend(["", "Flagged anomalies:"])
    if anomalies:
        lines.extend(f"- {item.get('severity', 'warning')}: {item.get('message', '')}" for item in anomalies)
    else:
        lines.append("- None.")
    return "\n".join(lines)


def _write_synthesis_summary(
    path: Path,
    validation: ReportValidation,
    lines: list[ReportLine],
    workbook_paths: list[Path],
    employee_name: str,
    title_prefix: str,
) -> None:
    payload = _summary_payload(validation, lines, workbook_paths, employee_name, title_prefix)
    summary = model_router.synthesize_report_summary(payload) or _fallback_summary_markdown(payload)
    path.write_text(summary + "\n", encoding="utf-8")


def _annotation_lines(lines: list[ReportLine]) -> list[ReceiptAnnotationLine]:
    return [
        ReceiptAnnotationLine(
            receipt_id=line.receipt_id,
            transaction_id=line.transaction_id,
            review_row_id=line.review_row_id,
            receipt_path=line.receipt_path,
            receipt_file_name=line.receipt_file_name,
            transaction_date=line.transaction_date,
            supplier=line.supplier,
            amount=line.amount,
            currency=line.currency,
            business_or_personal=line.business_or_personal,
            report_bucket=line.report_bucket,
            business_reason=line.business_reason,
            attendees=line.attendees,
        )
        for line in lines
    ]


def generate_report_package(
    session: Session,
    *,
    expense_report_id: int,
    employee_name: str,
    title_prefix: str,
    allow_warnings: bool = True,
) -> ReportRun:
    report = session.get(ExpenseReport, expense_report_id)
    if report is None:
        raise ValueError(f"ExpenseReport {expense_report_id} not found")
    if report.report_kind == "personal_reimbursement":
        raise NotImplementedError(
            "Personal reimbursement report template coming in M1 Day 8-9"
        )
    if report.report_kind != "diners_statement":
        raise ValueError(f"Unknown report_kind: {report.report_kind}")
    if report.statement_import_id is None:
        raise ValueError(
            f"Diners-statement report {expense_report_id} has no statement_import_id"
        )
    statement_import_id = report.statement_import_id

    settings = get_settings()
    if not settings.report_template_path or not settings.report_template_path.exists():
        raise FileNotFoundError("Expense report template was not found. Set EXPENSE_REPORT_TEMPLATE_PATH.")

    validation = validate_report_readiness(session, expense_report_id=expense_report_id)
    if validation.issue_count:
        if any(issue.code == "review_not_confirmed" for issue in validation.issues):
            raise ValueError("Report generation requires confirmed review data")
        raise ValueError(f"Report has {validation.issue_count} blocking validation error(s)")
    if validation.warning_count and not allow_warnings:
        raise ValueError(f"Report has {validation.warning_count} validation warning(s)")

    lines = _confirmed_lines(session, expense_report_id=expense_report_id)
    if not lines:
        raise ValueError("No confirmed review rows are available for report generation")

    run = ReportRun(
        statement_import_id=statement_import_id,
        expense_report_id=expense_report_id,
        template_name=settings.report_template_path.name,
        status="running",
    )
    session.add(run)
    session.commit()
    session.refresh(run)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    output_dir = settings.storage_root / "reports" / f"report_{run.id}_{timestamp}"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Bug 1: period-ending uses BMO statement_date when present. Falls back
    # to max(transaction_date) so manual-entry statements (no statement_date)
    # still produce a sensible header.
    statement = session.get(StatementImport, statement_import_id)
    statement_date = statement.statement_date if statement is not None else None
    period_ending = _resolve_period_ending(statement_date, lines)

    dates = sorted({line.transaction_date for line in lines})
    chunks = [dates[i : i + 14] for i in range(0, len(dates), 14)]
    workbook_paths: list[Path] = []
    for idx, chunk_dates in enumerate(chunks, start=1):
        chunk_lines = [line for line in lines if line.transaction_date in set(chunk_dates)]
        title = f"{title_prefix} - Part {idx}" if len(chunks) > 1 else title_prefix
        workbook_path = output_dir / f"expense_report_part_{idx}.xlsx"
        _fill_workbook(
            settings.report_template_path,
            workbook_path,
            employee_name,
            title,
            chunk_lines,
            period_ending=period_ending,
        )
        workbook_paths.append(workbook_path)

    summary_path = output_dir / "validation_summary.txt"
    _write_summary(summary_path, validation, lines, workbook_paths)
    synthesis_summary_path = output_dir / "summary.md"
    _write_synthesis_summary(synthesis_summary_path, validation, lines, workbook_paths, employee_name, title_prefix)
    annotated_pdf_path = output_dir / "annotated_receipts.pdf"
    create_annotated_receipts_pdf(_annotation_lines(lines), annotated_pdf_path)

    if len(workbook_paths) == 1 and not annotated_pdf_path.exists():
        final_path = workbook_paths[0]
    else:
        final_path = output_dir / "expense_report_package.zip"
        with ZipFile(final_path, "w", compression=ZIP_DEFLATED) as zf:
            for workbook_path in workbook_paths:
                zf.write(workbook_path, workbook_path.name)
            zf.write(summary_path, summary_path.name)
            zf.write(synthesis_summary_path, synthesis_summary_path.name)
            zf.write(annotated_pdf_path, annotated_pdf_path.name)

    run.status = "completed"
    run.output_workbook_path = str(final_path)
    run.output_pdf_path = str(annotated_pdf_path)
    session.add(run)
    session.commit()
    session.refresh(run)
    return run
