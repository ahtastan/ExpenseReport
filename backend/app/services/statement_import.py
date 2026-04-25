from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import re

from openpyxl import load_workbook
from sqlmodel import Session

from app.models import StatementImport, StatementTransaction

_AMOUNT_QUANT = Decimal("0.0001")


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m.%d.%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _swap_month_day(value: date) -> date | None:
    try:
        return date(value.year, value.day, value.month)
    except ValueError:
        return None


def _repair_swapped_excel_date_outliers(records: list[dict[str, Any]]) -> None:
    text_dates = [record["transaction_date"] for record in records if record["transaction_date"] and not record["date_from_excel"]]
    if not text_dates:
        return
    lower = min(text_dates) - timedelta(days=3)
    upper = max(text_dates) + timedelta(days=14)
    for record in records:
        tx_date = record["transaction_date"]
        if not tx_date or not record["date_from_excel"]:
            continue
        swapped = _swap_month_day(tx_date)
        if swapped and lower <= swapped <= upper and not (lower <= tx_date <= upper):
            record["transaction_date"] = swapped


def _parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        # str() avoids float-binary noise (Decimal(0.1) != Decimal("0.1")).
        return Decimal(str(value)).quantize(_AMOUNT_QUANT)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("TRY", "").replace("USD", "").replace("$", "").replace(" ", "")
    text = text.replace(",", "")
    try:
        return Decimal(text).quantize(_AMOUNT_QUANT)
    except (InvalidOperation, ValueError):
        return None


def _normalize_supplier(value: Any) -> str:
    text = str(value or "").strip()
    return " ".join(text.upper().split())


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _header_index(headers: list[str], candidates: set[str]) -> int | None:
    normalized = [_normalize_header(h) for h in headers]
    normalized_candidates = {_normalize_header(candidate) for candidate in candidates}
    for idx, header in enumerate(normalized):
        if header in normalized_candidates:
            return idx
    return None


TRANSACTION_DATE_HEADERS = {
    "tran date",
    "trans date",
    "transaction date",
    "txn date",
    "date",
}
SUPPLIER_HEADERS = {
    "supplier",
    "supplier name",
    "merchant",
    "merchant name",
    "description",
    "vendor",
    "vendor name",
}
SOURCE_AMOUNT_HEADERS = {
    "source amount",
    "local amount",
    "amount local",
    "transaction amount",
    "amount",
}
USD_AMOUNT_HEADERS = {
    "amount incl usd",
    "amount usd",
    "usd amount",
    "amount incl",
}


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, int, int, int | None, int | None]:
    best_found: set[str] = set()
    for row_idx, row in enumerate(rows[:30]):
        headers = [str(value or "").strip() for value in row]
        tran_idx = _header_index(headers, TRANSACTION_DATE_HEADERS)
        supplier_idx = _header_index(headers, SUPPLIER_HEADERS)
        source_amount_idx = _header_index(headers, SOURCE_AMOUNT_HEADERS)
        usd_idx = _header_index(headers, USD_AMOUNT_HEADERS)
        found = set()
        if tran_idx is not None:
            found.add("transaction date")
        if supplier_idx is not None:
            found.add("supplier")
        if len(found) > len(best_found):
            best_found = found
        if tran_idx is not None and supplier_idx is not None:
            return row_idx, tran_idx, supplier_idx, source_amount_idx, usd_idx

    missing = [name for name in ("transaction date", "supplier") if name not in best_found]
    if not missing:
        missing = ["transaction date and supplier in the same header row"]
    raise ValueError(f"Could not find required statement columns: {', '.join(missing)}")


def import_diners_excel(
    session: Session,
    file_path: Path,
    source_filename: str,
    uploader_user_id: int | None = None,
) -> StatementImport:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty")

    header_row_idx, tran_idx, supplier_idx, source_amount_idx, usd_idx = _find_header_row(rows)

    statement = StatementImport(
        uploader_user_id=uploader_user_id,
        source_filename=source_filename,
        storage_path=str(file_path),
    )
    session.add(statement)
    session.commit()
    session.refresh(statement)

    pending_transactions: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not any(value not in (None, "") for value in row):
            continue
        supplier_raw = str(row[supplier_idx] or "").strip()
        if not supplier_raw:
            continue
        raw_date = row[tran_idx]
        transaction_date = _parse_date(raw_date)
        local_amount = _parse_amount(row[source_amount_idx]) if source_amount_idx is not None else None
        usd_amount = _parse_amount(row[usd_idx]) if usd_idx is not None else None
        pending_transactions.append(
            {
                "transaction_date": transaction_date,
                "date_from_excel": isinstance(raw_date, (datetime, date)),
                "supplier_raw": supplier_raw,
                "local_amount": local_amount,
                "usd_amount": usd_amount,
                "source_row_ref": str(row_number),
            }
        )

    _repair_swapped_excel_date_outliers(pending_transactions)

    transaction_count = 0
    min_date: date | None = None
    max_date: date | None = None
    for pending in pending_transactions:
        transaction_date = pending["transaction_date"]
        if transaction_date:
            min_date = transaction_date if min_date is None else min(min_date, transaction_date)
            max_date = transaction_date if max_date is None else max(max_date, transaction_date)
        transaction = StatementTransaction(
            statement_import_id=statement.id,
            transaction_date=transaction_date,
            supplier_raw=pending["supplier_raw"],
            supplier_normalized=_normalize_supplier(pending["supplier_raw"]),
            local_currency="TRY",
            local_amount=pending["local_amount"],
            usd_amount=pending["usd_amount"],
            source_row_ref=pending["source_row_ref"],
            source_kind="excel",
        )
        session.add(transaction)
        transaction_count += 1

    statement.row_count = transaction_count
    statement.period_start = min_date
    statement.period_end = max_date
    session.add(statement)
    session.commit()
    session.refresh(statement)
    wb.close()
    return statement
