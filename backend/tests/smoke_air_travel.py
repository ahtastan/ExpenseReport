"""Smoke test for Air Travel Reconciliation end-to-end wiring.

Verifies:
- New review rows get air_travel_* defaults via `_default_air_travel`.
- `update_review_row` accepts air_travel_* fields through to confirmed_json.
- `_confirmed_lines` surfaces the values on `ReportLine`.
- `_fill_workbook` writes row 47 on Week 1A with the detail fields and
  leaves column J (formula `=H-I`) untouched.
"""

import os
from datetime import date
from pathlib import Path
from uuid import uuid4

VERIFY_ROOT = Path.cwd() / ".verify_data"
VERIFY_ROOT.mkdir(parents=True, exist_ok=True)
os.environ["DATABASE_URL"] = f"sqlite:///{VERIFY_ROOT / f'air_travel_{uuid4().hex}.db'}"
os.environ["EXPENSE_STORAGE_ROOT"] = str(VERIFY_ROOT)
_default_template = Path.cwd().parent / "Expense Report Form_Blank.xlsx"
if not _default_template.exists():
    _default_template = Path.cwd().parent.parent / "Expense Report Form_Blank.xlsx"
os.environ["EXPENSE_REPORT_TEMPLATE_PATH"] = str(_default_template)
os.environ.pop("ANTHROPIC_API_KEY", None)

from sqlmodel import Session  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.db import create_db_and_tables, engine  # noqa: E402
from app.models import StatementImport, StatementTransaction  # noqa: E402
from app.services.report_generator import (  # noqa: E402
    AIRFARE_BUCKET,
    AIR_TRAVEL_ROWS_BY_SHEET,
    _confirmed_lines,
    _fill_workbook,
)
from app.services.report_validation import validate_report_readiness  # noqa: E402
from app.services.review_sessions import (  # noqa: E402
    confirm_review_session,
    get_or_create_review_session,
    review_rows,
    update_review_row,
)


def main() -> None:
    create_db_and_tables()
    with Session(engine) as session:
        imp = StatementImport(source_filename="air_travel_smoke.xlsx", storage_path="(memory)")
        session.add(imp)
        session.commit()
        session.refresh(imp)

        # One airfare transaction (populates air travel detail row 47 on Week 1A)
        tx1 = StatementTransaction(
            statement_import_id=imp.id,
            transaction_date=date(2026, 3, 12),
            supplier_raw="PEGASUS HAVA YOLLARI",
            supplier_normalized="pegasus hava yollari",
            local_amount=0,
            local_currency="USD",
            usd_amount=0,
            source_row_ref="row-1",
        )
        # One non-airfare transaction to confirm the air-travel filler skips it
        tx2 = StatementTransaction(
            statement_import_id=imp.id,
            transaction_date=date(2026, 3, 13),
            supplier_raw="HILTON ISTANBUL",
            supplier_normalized="hilton istanbul",
            local_amount=812.34,
            local_currency="USD",
            usd_amount=812.34,
            source_row_ref="row-2",
        )
        session.add_all([tx1, tx2])
        session.commit()

        review = get_or_create_review_session(session, imp.id)
        rows = review_rows(session, review.id)
        assert len(rows) == 2, f"expected 2 rows, got {len(rows)}"

        # Verify air travel defaults are present on the fresh row
        import json

        row_air = next(r for r in rows if "PEGASUS" in (json.loads(r.suggested_json)["supplier"] or ""))
        confirmed = json.loads(row_air.confirmed_json)
        assert confirmed.get("air_travel_paid_by") == "DC Card"
        assert "air_travel_date" in confirmed
        print("defaults: OK  air_travel_paid_by=DC Card; keys present")

        # Patch the airfare row with required + air-travel detail fields
        update_review_row(
            session,
            row_id=row_air.id,
            fields={
                "business_or_personal": "Business",
                "report_bucket": AIRFARE_BUCKET,
                "business_reason": "Kartonsan visit",
                "air_travel_date": "2026-03-12",
                "air_travel_from": "IST",
                "air_travel_to": "ESB",
                "air_travel_airline": "Pegasus",
                "air_travel_rt_or_oneway": "RT",
                "air_travel_return_date": "2026-03-15",
                "air_travel_paid_by": "DC Card",
                "air_travel_total_tkt_cost": 523.45,
                "air_travel_prior_tkt_value": 0,
                "air_travel_comments": "Smoke test",
            },
        )

        # Patch the non-airfare row with required fields (must be non-flagged to confirm)
        row_hotel = next(r for r in rows if r.id != row_air.id)
        update_review_row(
            session,
            row_id=row_hotel.id,
            fields={
                "business_or_personal": "Business",
                "report_bucket": "Hotel/Lodging/Laundry",
                "business_reason": "Kartonsan visit",
            },
        )

        confirm_review_session(session, review.id, confirmed_by_label="air-travel-smoke")

        # _confirmed_lines should carry the air travel fields onto ReportLine
        lines = _confirmed_lines(session, imp.id)
        assert len(lines) == 2
        air_line = next(ln for ln in lines if ln.report_bucket == AIRFARE_BUCKET)
        assert air_line.air_travel_from == "IST"
        assert air_line.air_travel_to == "ESB"
        assert air_line.air_travel_airline == "Pegasus"
        assert air_line.air_travel_rt_or_oneway == "RT"
        assert air_line.air_travel_return_date == date(2026, 3, 15)
        assert air_line.air_travel_paid_by == "DC Card"
        assert air_line.amount == 0
        assert air_line.air_travel_total_tkt_cost == 523.45
        assert air_line.air_travel_prior_tkt_value == 0
        assert air_line.air_travel_comments == "Smoke test"
        print("confirmed_lines: OK  ReportLine populated with air travel fields")

        # Render into a workbook and verify cells
        template_path = Path(os.environ["EXPENSE_REPORT_TEMPLATE_PATH"])
        if not template_path.exists():
            print(f"SKIP workbook render: template not found at {template_path}")
            return
        output_path = VERIFY_ROOT / f"air_travel_smoke_{uuid4().hex}.xlsx"
        _fill_workbook(template_path, output_path, "Smoke Tester", "Air Travel Smoke", lines)

        wb = load_workbook(output_path)
        ws = wb["Week 1A"]
        row47 = AIR_TRAVEL_ROWS_BY_SHEET["Week 1A"][0]
        print(
            f"Week 1A row {row47}:",
            f"B={ws[f'B{row47}'].value}",
            f"C={ws[f'C{row47}'].value}",
            f"D={ws[f'D{row47}'].value}",
            f"E={ws[f'E{row47}'].value}",
            f"F={ws[f'F{row47}'].value}",
            f"G={ws[f'G{row47}'].value}",
            f"H={ws[f'H{row47}'].value}",
            f"I={ws[f'I{row47}'].value}",
            f"J={ws[f'J{row47}'].value}  (formula expected)",
            f"K={ws[f'K{row47}'].value}",
        )
        assert ws[f"B{row47}"].value == "12.03.2026 - 15.03.2026", ws[f"B{row47}"].value
        assert ws[f"C{row47}"].value == "IST", ws[f"C{row47}"].value
        assert ws[f"D{row47}"].value == "ESB"
        assert ws[f"E{row47}"].value == "Pegasus"
        assert ws[f"F{row47}"].value == "RT"
        assert ws[f"G{row47}"].value == "DC Card"
        assert ws[f"H{row47}"].value == 523.45
        assert ws[f"I{row47}"].value == 0
        assert ws[f"K{row47}"].value == "Smoke test"
        # Column J must not be overwritten with a scalar; it should remain a formula or None.
        j_val = ws[f"J{row47}"].value
        assert j_val is None or (isinstance(j_val, str) and j_val.startswith("=")), (
            f"Column J was overwritten with scalar: {j_val!r}"
        )
        print("workbook: OK  row 47 populated, column J preserved")

        assert ws["E7"].value == 523.45, ws["E7"].value
        print("workbook: OK  main airfare row populated from ticket cost on transaction date")

        # Confirm row 48 (next unused) is empty on this sheet
        row48 = AIR_TRAVEL_ROWS_BY_SHEET["Week 1A"][1]
        assert ws[f"C{row48}"].value in (None, ""), ws[f"C{row48}"].value
        print(f"workbook: OK  Week 1A row {row48} left blank")

        overflow_imp = StatementImport(source_filename="air_travel_overflow.xlsx", storage_path="(memory)")
        session.add(overflow_imp)
        session.commit()
        session.refresh(overflow_imp)
        for index in range(4):
            session.add(
                StatementTransaction(
                    statement_import_id=overflow_imp.id,
                    transaction_date=date(2026, 3, 14 + index),
                    supplier_raw=f"AIRLINE {index + 1}",
                    supplier_normalized=f"airline {index + 1}",
                    local_amount=100 + index,
                    local_currency="USD",
                    usd_amount=100 + index,
                    source_row_ref=f"overflow-{index + 1}",
                )
            )
        session.commit()
        overflow_review = get_or_create_review_session(session, overflow_imp.id)
        for row in review_rows(session, overflow_review.id):
            update_review_row(
                session,
                row_id=row.id,
                fields={
                    "business_or_personal": "Business",
                    "report_bucket": AIRFARE_BUCKET,
                    "air_travel_total_tkt_cost": 100,
                },
            )
        confirm_review_session(session, overflow_review.id, confirmed_by_label="air-overflow-smoke")
        validation = validate_report_readiness(session, overflow_imp.id)
        assert validation.warning_count == 1
        assert validation.issues[0].code == "air_travel_detail_overflow"
        assert "Week 1A" in validation.issues[0].message
        print("validation: OK  air travel overflow warning emitted")

        missing_return_imp = StatementImport(source_filename="air_travel_missing_return.xlsx", storage_path="(memory)")
        session.add(missing_return_imp)
        session.commit()
        session.refresh(missing_return_imp)
        session.add(
            StatementTransaction(
                statement_import_id=missing_return_imp.id,
                transaction_date=date(2026, 3, 20),
                supplier_raw="ROUNDTRIP AIR",
                supplier_normalized="roundtrip air",
                local_amount=250,
                local_currency="USD",
                usd_amount=250,
                source_row_ref="missing-return-1",
            )
        )
        session.commit()
        missing_return_review = get_or_create_review_session(session, missing_return_imp.id)
        missing_return_row = review_rows(session, missing_return_review.id)[0]
        update_review_row(
            session,
            row_id=missing_return_row.id,
            fields={
                "business_or_personal": "Business",
                "report_bucket": AIRFARE_BUCKET,
                "air_travel_rt_or_oneway": "RT",
                "air_travel_total_tkt_cost": 250,
            },
        )
        confirm_review_session(session, missing_return_review.id, confirmed_by_label="air-missing-return-smoke")
        missing_return_validation = validate_report_readiness(session, missing_return_imp.id)
        assert missing_return_validation.issue_count == 1
        missing_issue = missing_return_validation.issues[0]
        assert missing_issue.code == "air_travel_return_date_missing"
        assert missing_issue.review_row_id == missing_return_row.id
        assert missing_issue.supplier == "ROUNDTRIP AIR"
        assert missing_issue.transaction_date == "2026-03-20"
        assert missing_issue.report_bucket == AIRFARE_BUCKET
        print("validation: OK  RT air travel requires return date")

        early_return_imp = StatementImport(source_filename="air_travel_early_return.xlsx", storage_path="(memory)")
        session.add(early_return_imp)
        session.commit()
        session.refresh(early_return_imp)
        session.add(
            StatementTransaction(
                statement_import_id=early_return_imp.id,
                transaction_date=date(2026, 3, 20),
                supplier_raw="EARLY RETURN AIR",
                supplier_normalized="early return air",
                local_amount=275,
                local_currency="USD",
                usd_amount=275,
                source_row_ref="early-return-1",
            )
        )
        session.commit()
        early_return_review = get_or_create_review_session(session, early_return_imp.id)
        early_return_row = review_rows(session, early_return_review.id)[0]
        update_review_row(
            session,
            row_id=early_return_row.id,
            fields={
                "business_or_personal": "Business",
                "report_bucket": AIRFARE_BUCKET,
                "air_travel_date": "2026-03-20",
                "air_travel_rt_or_oneway": "RT",
                "air_travel_return_date": "2026-03-19",
                "air_travel_total_tkt_cost": 275,
            },
        )
        confirm_review_session(session, early_return_review.id, confirmed_by_label="air-early-return-smoke")
        early_return_validation = validate_report_readiness(session, early_return_imp.id)
        assert early_return_validation.issue_count == 1
        early_issue = early_return_validation.issues[0]
        assert early_issue.code == "air_travel_return_date_before_travel_date"
        assert early_issue.review_row_id == early_return_row.id
        assert early_issue.supplier == "EARLY RETURN AIR"
        assert early_issue.transaction_date == "2026-03-20"
        assert early_issue.report_bucket == AIRFARE_BUCKET
        print("validation: OK  RT return date cannot be before travel date")

        print("AIR TRAVEL SMOKE TEST PASSED")


if __name__ == "__main__":
    main()
