"""Bug 3: Page 1B IRS detail cells must NOT hard-truncate place/location/
participants/reason text. Long values (a 5-person attendee list, a
multi-clause business reason) must flow vertically via wrap_text.

Symptom on the November 2025 demo: FERMAKI participants showed
``"Ahmet Hakan Taştan, Burak Ibar, Kaspar "`` (cut mid-name at 40 chars).
Business reason: ``"EDT team dinner Izmir — internal team meal during"``
(cut mid-sentence at 50 chars). The slices were ``primary.place[:40]``,
``primary.location[:28]``, ``primary.participants[:40]``,
``primary.reason[:50]``.

Fix: write full strings, apply ``Alignment(wrap_text=True, vertical='top')``
to the four wide-content cells, set ``row_dimensions[rownum].height = None``
so Excel auto-fits.
"""

from __future__ import annotations

import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services.report_generator import (  # noqa: E402
    MealDetailLine,
    _write_irs_detail_row,
)


def _fermaki_detail() -> MealDetailLine:
    """Realistic detail with attendees + reason long enough to trigger the
    old [:40] / [:50] slice."""
    return MealDetailLine(
        tx_date=date(2025, 10, 20),
        code="D",
        place="Fermaki Meat & More / Salen Mirrestoran",
        location="Izmir",
        participants="Ahmet Hakan Taştan, Burak Ibar, Kaspar Yıldız, Erdem Ç.",
        reason=(
            "EDT team dinner Izmir — internal team meal during customer "
            "site visit, 4 EDT attendees (split bill 1 of 2, total 4465 TL)"
        ),
        amount=Decimal("92.72"),
        eg=False,
        mr=True,
    )


def test_irs_row_writes_full_attendees_without_slice() -> None:
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, None)

    # Full attendees written to E31 — no [:40] slice.
    assert ws["E31"].value == primary.participants
    assert len(ws["E31"].value) > 40, (
        "fixture must be longer than the old slice limit to verify the bug"
    )


def test_irs_row_writes_full_business_reason_without_slice() -> None:
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, None)

    assert ws["F31"].value == primary.reason
    assert len(ws["F31"].value) > 50, (
        "fixture must be longer than the old slice limit to verify the bug"
    )


def test_irs_row_applies_wrap_text_to_wide_content_cells() -> None:
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, None)

    for col in ("C", "D", "E", "F"):
        cell = ws[f"{col}31"]
        assert cell.alignment.wrap_text is True, (
            f"cell {col}31 must have wrap_text=True for long text to render"
        )
        assert cell.alignment.vertical == "top", (
            f"cell {col}31 should be top-aligned so wrapped text reads naturally"
        )


def test_irs_row_height_is_auto() -> None:
    """row_dimensions[r].height = None so Excel auto-fits on render."""
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, None)

    # openpyxl: height=None means "use Excel's default auto behavior".
    assert ws.row_dimensions[31].height is None


def test_irs_row_writes_sum_formula_when_sum_components_present() -> None:
    """Bug 2 + Bug 3 interaction: helper still writes the SUM formula to J
    column when sum_components carries multiple amounts."""
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, [Decimal("92.72"), Decimal("16.65")])

    assert ws["J31"].value == "=92.72+16.65", (
        f"expected SUM formula in J31, got {ws['J31'].value!r}"
    )


def test_irs_row_leaves_amount_cell_alone_when_no_sum_components() -> None:
    """Single-receipt case: J cell stays None so the template's pre-existing
    IF formula isn't overridden (caller relies on the template formula
    pulling from Week 1A daily totals)."""
    wb = Workbook()
    ws = wb.active
    primary = _fermaki_detail()

    _write_irs_detail_row(ws, 31, primary, None)

    assert ws["J31"].value is None, (
        f"J31 should be untouched when sum_components is None; got {ws['J31'].value!r}"
    )
