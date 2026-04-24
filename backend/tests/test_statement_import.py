import asyncio
import os
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

VERIFY_ROOT = Path.cwd() / ".verify_data"
VERIFY_ROOT.mkdir(parents=True, exist_ok=True)
os.environ["DATABASE_URL"] = f"sqlite:///{VERIFY_ROOT / f'statement_import_{uuid4().hex}.db'}"
os.environ["EXPENSE_STORAGE_ROOT"] = str(VERIFY_ROOT)

from fastapi import HTTPException, UploadFile  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from sqlmodel import Session, select  # noqa: E402

from app.db import create_db_and_tables, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import AppUser, StatementImport  # noqa: E402
from app.routes.statements import import_statement_excel  # noqa: E402
from app.services.statement_import import import_diners_excel  # noqa: E402


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def upload_from_workbook(rows: list[list[object]], filename: str = "statement.xlsx") -> UploadFile:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return UploadFile(file=buffer, filename=filename)


def workbook_bytes(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def import_fixture(rows: list[list[object]], filename: str) -> tuple[int, date | None, date | None]:
    path = VERIFY_ROOT / filename
    write_workbook(path, rows)
    with Session(engine) as session:
        statement = import_diners_excel(session, path, filename)
        return statement.row_count, statement.period_start, statement.period_end


def test_import_finds_header_after_metadata_rows() -> None:
    row_count, period_start, period_end = import_fixture(
        [
            ["Card Transactions"],
            ["03/12/2026 to 04/11/2026"],
            ["Tastan Hakan - Diners Club - (7208)"],
            [],
            ["Tran Date", "Supplier", "Source Amount", "Amount Tax", "Amount Incl"],
            ["03/13/2026", "Faturamati Taksi Yaz", "562.50 TRY", 0, 13.07],
            ["03/14/2026", "Takside Pos", "439.56 TRY", 0, 10.22],
        ],
        "metadata_header.xlsx",
    )
    assert row_count == 2
    assert period_start == date(2026, 3, 13)
    assert period_end == date(2026, 3, 14)


def test_import_tolerates_header_spacing_and_name_variants() -> None:
    row_count, period_start, period_end = import_fixture(
        [
            ["TRANSACTION  DATE", "Supplier Name", "Source  Amount"],
            ["03/15/2026", "Param Param Yemeksepet", "2,383.29 TRY"],
        ],
        "variant_headers.xlsx",
    )
    assert row_count == 1
    assert period_start == date(2026, 3, 15)
    assert period_end == date(2026, 3, 15)


def test_import_treats_diners_dates_as_month_first() -> None:
    row_count, period_start, period_end = import_fixture(
        [
            ["Tran Date", "Supplier", "Source Amount", "Amount Incl"],
            ["04/01/2026", "Aat Istanbul Airport S", "550.00 TRY", 12.68],
            ["04/08/2026", "Uber Trip", "212.10 TRY", 4.88],
        ],
        "month_first_dates.xlsx",
    )
    assert row_count == 2
    assert period_start == date(2026, 4, 1)
    assert period_end == date(2026, 4, 8)


def test_import_repairs_swapped_excel_date_outliers() -> None:
    row_count, period_start, period_end = import_fixture(
        [
            ["Tran Date", "Supplier", "Source Amount", "Amount Incl"],
            [datetime(2026, 11, 3), "Faturamati Taksi Yaz", "394.88 TRY", 9.25],
            [datetime(2026, 12, 3), "Takside Pos", "444.00 TRY", 10.32],
            ["03/13/2026", "Uber Trip", "387.70 TRY", 9.01],
            ["04/08/2026", "Volkan Market", "290.00 TRY", 6.68],
        ],
        "swapped_excel_outliers.xlsx",
    )
    assert row_count == 4
    assert period_start == date(2026, 3, 11)
    assert period_end == date(2026, 4, 8)


def test_real_diners_transactions_fixture_imports() -> None:
    path = Path(r"C:/Users/CASPER/.openclaw/workspace/Expense/03_11_Receipts/Diners_Transactions.xlsx")
    if not path.exists():
        raise AssertionError(f"Expected real Diners fixture at {path}")
    with Session(engine) as session:
        statement = import_diners_excel(session, path, path.name)
        assert statement.row_count > 0
        assert statement.period_start == date(2026, 3, 11)
        assert statement.period_end == date(2026, 4, 8)


def test_import_route_returns_400_for_missing_required_columns() -> None:
    upload = upload_from_workbook(
        [
            ["Card Transactions"],
            ["Amount", "Notes"],
            [123.45, "missing required headers"],
        ],
        "missing_headers.xlsx",
    )
    with Session(engine) as session:
        try:
            asyncio.run(import_statement_excel(file=upload, session=session))
            raise AssertionError("Expected missing headers to become a client error")
        except HTTPException as exc:
            assert exc.status_code == 400
            assert "transaction date" in str(exc.detail)
            assert "supplier" in str(exc.detail)


def test_browser_import_assigns_stable_owner_and_unblocks_review_and_validation() -> None:
    client = TestClient(app)
    payload = workbook_bytes(
        [
            ["Tran Date", "Supplier", "Source Amount", "Amount Incl"],
            ["04/01/2026", "Smoke Import Market", "123.45 TRY", 2.85],
        ]
    )

    first = client.post(
        "/statements/import-excel",
        files={
            "file": (
                "browser_statement.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    second = client.post(
        "/statements/import-excel",
        files={
            "file": (
                "browser_statement_again.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert first.status_code == 200, first.text
    assert second.status_code == 200, second.text
    first_body = first.json()
    second_body = second.json()
    assert first_body["uploader_user_id"] is not None
    assert second_body["uploader_user_id"] == first_body["uploader_user_id"]

    review = client.get(f"/reviews/report/{first_body['id']}")
    assert review.status_code == 200, review.text
    assert review.json()["rows"]

    validation = client.get(f"/reports/validate/{first_body['id']}")
    assert validation.status_code != 422, validation.text
    assert "Statement has no uploader" not in validation.text

    with Session(engine) as session:
        demo_users = session.exec(
            select(AppUser).where(AppUser.username == "demo")
        ).all()
    assert len(demo_users) == 1


def test_manual_statement_creation_assigns_owner_when_creating_statement() -> None:
    client = TestClient(app)

    response = client.post(
        "/statements/manual/transactions",
        json={
            "transaction_date": "2026-04-02",
            "supplier": "Manual Demo Market",
            "amount": 42.5,
            "currency": "TRY",
        },
    )

    assert response.status_code == 200, response.text
    review_session = response.json()["review_session"]
    statement_id = review_session["statement_import_id"]

    with Session(engine) as session:
        statement = session.get(StatementImport, statement_id)
        assert statement is not None
        assert statement.uploader_user_id is not None

    review = client.get(f"/reviews/report/{statement_id}")
    assert review.status_code == 200, review.text


def main() -> None:
    create_db_and_tables()
    test_import_finds_header_after_metadata_rows()
    test_import_tolerates_header_spacing_and_name_variants()
    test_import_treats_diners_dates_as_month_first()
    test_import_repairs_swapped_excel_date_outliers()
    test_real_diners_transactions_fixture_imports()
    test_import_route_returns_400_for_missing_required_columns()
    print("statement_import_tests=passed")


if __name__ == "__main__":
    main()
